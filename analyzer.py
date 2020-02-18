# encoding: utf-8
import datetime
import os
import openpyxl
import matplotlib.pyplot as pl
import numpy


DAILY_REPORT_IN_EXCEL_PATH = './huanggang/'


def extract_data_from_official_daily_report_in_excel():
	now = datetime.datetime.now()
	day = now - datetime.timedelta(days=30)
	case_list = list()
	date_list = list()

	while day <= now:
		try:
			str_day = datetime.datetime.strftime(day, '%Y%m%d')
			report_file = DAILY_REPORT_IN_EXCEL_PATH + str_day + '.xlsx'
			if os.path.exists(report_file):
				date_list.append(str_day)
				work_book = openpyxl.load_workbook(report_file)
				sheet_obj = work_book.active

				case = dict()
				region = dict()

				i = 3
				while i <= 13:
					# Start Feb. 13, data of '临床诊断病例' was added in column 5.
					confirmed_cases = sheet_obj.cell(row=i, column=5).value
					if not confirmed_cases:
						confirmed_cases = sheet_obj.cell(row=i, column=2).value
					else:
						confirmed_cases += sheet_obj.cell(row=i, column=2).value

					region[sheet_obj.cell(row=i, column=1).value] = {
						'confirmed': confirmed_cases,
						'cured': sheet_obj.cell(row=i, column=3).value,
						'dead': sheet_obj.cell(row=i, column=4).value,
					}
					i += 1

				case[str_day] = {
					'newly_added': region
				}

				case_list.append(case)

			day += datetime.timedelta(days=1)
		except Exception as e:
			print(e)
	return date_list, case_list
	print(F'date_list: {date_list}. case_list: {case_list}')


# def mock_early_case(end_date, str_start_date='20200119'):
# 	dates = list()
# 	newly_added_confirmed_cases = list()
# 	newly_added_cured_cases = list()
# 	newly_added_dead_cases = list()
# 	start_date = datetime.datetime.strptime(str_start_date, '%Y%m%d')
# 	n = (end_date - start_date).days
# 	for i in range(n):
# 		dates.append(datetime.datetime.strftime(start_date + datetime.timedelta(days=i), '%Y%m%d'))
# 		if i == n - 1:
# 			newly_added_confirmed_cases.append(64) # 122 - 58, data from official report of Jan 25
# 		else:
# 			newly_added_confirmed_cases.append(0)
# 		newly_added_cured_cases.append(0)
# 		newly_added_dead_cases.append(0)
# 	return dates, newly_added_confirmed_cases, newly_added_cured_cases, newly_added_dead_cases


def draw_daily_case_figure(date_list, case_number_list, title='疫情新增趋势图', city='', color='red',
						   case_number_list2=None, color2='red'):
	title = city + title
	pl.rcParams['font.family'] = 'sans-serif'
	pl.rcParams['font.serif'] = ['Heiti']
	pl.rcParams["figure.figsize"] = (8, 4)

	pl.xticks(rotation=70)
	# pl.plot(date_list, case_number_list, 'r', markevery=100)
	pl.plot(date_list, case_number_list, color=color, marker='o', linestyle='-', markersize=6)

	if case_number_list2:
		pl.plot(date_list, case_number_list2, color=color2, marker='o', linestyle='-', markersize=6)

	pl.grid(color='grey', axis='y')
	# pl.scatter(date_list, case_number_list)
	pl.title(title)

	# pl.show()
	now = datetime.datetime.now()
	folder_name = datetime.datetime.strftime(now, '%Y%m%d')
	folder_path = F'reports/{folder_name}/{city}'
	if not os.path.exists(folder_path):
		os.mkdir(folder_path)
	t = datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d-%H%M%S')
	pl.savefig(F'{folder_path}/{title}-{t}.png')
	pl.close()


def draw_bar_figure_by_all_regions(date_list, case_number_list, title='黄冈各县市疫情确诊累计柱状图'):
	pl.rcParams['font.family'] = 'sans-serif'
	pl.rcParams['font.serif'] = ['Heiti']
	pl.rcParams["figure.figsize"] = (8, 4)

	# pl.plot(date_list, case_number_list, 'r', markevery=100)
	pl.bar(date_list, case_number_list, color='red')
	pl.grid(color='grey', axis='y')
	# pl.scatter(date_list, case_number_list)
	pl.title(title)
	# pl.show()
	now = datetime.datetime.now()
	folder_name = datetime.datetime.strftime(now, '%Y%m%d')
	folder_path = F'reports/{folder_name}'
	if not os.path.exists(folder_path):
		os.mkdir(folder_path)
	t = datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d-%H%M%S')
	pl.savefig(F'{folder_path}/{title}-{t}.png')
	pl.close()


def write_data_to_excel(date_list, case_number_list, excel_name='data.xlsx'):
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.cell(row=1, column=1, value='日期')
	ws.cell(row=2, column=1, value='人数')
	for col in range(1, len(date_list) + 1):
		ws.cell(row=1, column=col + 1, value=date_list[col-1])
		ws.cell(row=2, column=col + 1, value=case_number_list[col - 1])
	wb.save(excel_name)


def sum_daily_added_cases(newly_added_case_number_list):
	accumulated_cases = list()

	accumulated_cases.append(newly_added_case_number_list[0])
	i = 1
	while i < len(newly_added_case_number_list):
		accumulated_case = accumulated_cases[i-1] + newly_added_case_number_list[i]
		accumulated_cases.append(accumulated_case)
		i += 1
	return accumulated_cases


if __name__ == '__main__':
	date_list, case_list = extract_data_from_official_daily_report_in_excel()

	newly_added_cases_by_regions = list()

	for c in case_list:
		for i in c:
			newly_added_cases_by_regions.append(c.get(i).get('newly_added'))

			# newly_added_confirmed_cases.append(c.get(i).get('newly_added').get('confirmed'))
			# newly_added_cured_cases.append(c.get(i).get('newly_added').get('cured'))
			# newly_added_dead_cases.append(c.get(i).get('newly_added').get('dead'))

	all_regions = newly_added_cases_by_regions[0].keys()

	newly_added_cases_dict = dict()

	for key in all_regions:
		newly_added_confirmed_cases = list()
		newly_added_cured_cases = list()
		newly_added_dead_cases = list()
		for i in newly_added_cases_by_regions:
			c = i[key]
			newly_added_confirmed_cases.append(c.get('confirmed'))
			newly_added_cured_cases.append(c.get('cured'))
			newly_added_dead_cases.append(c.get('dead'))
		newly_added_cases_dict[key] = {
			'confirmed': newly_added_confirmed_cases,
			'cured': newly_added_cured_cases,
			'dead': newly_added_dead_cases,
		}

	whole_city_newly_added_confirmed_cases = newly_added_cases_dict['全市累计']['confirmed']
	whole_city_newly_cured_confirmed_cases = newly_added_cases_dict['全市累计']['cured']
	whole_city_newly_dead_confirmed_cases = newly_added_cases_dict['全市累计']['dead']

	whole_city_accumulated_confirmed_case_list = sum_daily_added_cases(whole_city_newly_added_confirmed_cases)
	whole_city_accumulated_added_cured_cases = sum_daily_added_cases(whole_city_newly_cured_confirmed_cases)
	whole_city_accumulated_added_dead_cases = sum_daily_added_cases(whole_city_newly_dead_confirmed_cases)

	simplified_date_list = list(map(lambda d: d.split('2020')[1], date_list))
	# As the x-ray is so crowed, so simplify the dates
	# simplified_date_list = list()
	# remove_tags = ('202001', '202002')
	# for d in date_list:
	# 	if remove_tags[0] in d:
	# 		t = d.split(remove_tags[0])[1]
	# 	elif remove_tags[1] in d:
	# 		t = '2.' + d.split(remove_tags[1])[1]
	# 	simplified_date_list.append(t)

	write_data_to_excel(simplified_date_list, whole_city_newly_added_confirmed_cases)

	accumulated_confirmed_case_list_by_regions = list()
	for n in newly_added_cases_dict:
		accumulated_confirmed_case_list_by_regions.append(sum_daily_added_cases(newly_added_cases_dict[n]['confirmed']).pop())
	region_list = list(all_regions)
	region_list.pop()
	accumulated_confirmed_case_list_by_regions.pop()

	zipped = zip(region_list, accumulated_confirmed_case_list_by_regions)
	temp = sorted(zipped, key=lambda x: x[1], reverse=True)
	sorted_region_list, sorted_accumulated_confirmed_case_list_by_regions = zip(*temp)

	draw_bar_figure_by_all_regions(sorted_region_list, sorted_accumulated_confirmed_case_list_by_regions,
								   F'黄冈各县市疫情确诊累计柱状图（截止到 {datetime.datetime.strftime(datetime.datetime.now() - datetime.timedelta(days=1), "%Y%m%d")}）')


	draw_daily_case_figure(simplified_date_list, whole_city_newly_added_confirmed_cases, '黄冈全市疫情新增趋势图')

	draw_daily_case_figure(simplified_date_list, whole_city_accumulated_confirmed_case_list, '黄冈全市疫情确诊累计趋势图')
	draw_daily_case_figure(simplified_date_list, whole_city_accumulated_added_cured_cases,
						   '黄冈全市疫情治愈(绿)-死亡(红)累计趋势图', city='', color='green',
						   case_number_list2=whole_city_accumulated_added_dead_cases, color2='red')


	# macheng
	target_city = '麻城'
	for city in all_regions:
		if city != target_city:
			continue
		city_newly_added_confirmed_cases = newly_added_cases_dict[city]['confirmed']
		city_newly_cured_confirmed_cases = newly_added_cases_dict[city]['cured']
		city_newly_dead_confirmed_cases = newly_added_cases_dict[city]['dead']

		city_accumulated_confirmed_case_list = sum_daily_added_cases(city_newly_added_confirmed_cases)
		city_accumulated_added_cured_cases = sum_daily_added_cases(city_newly_cured_confirmed_cases)
		city_accumulated_added_dead_cases = sum_daily_added_cases(city_newly_dead_confirmed_cases)

		draw_daily_case_figure(simplified_date_list, city_newly_added_confirmed_cases, '疫情新增趋势图', city)

		draw_daily_case_figure(simplified_date_list, city_accumulated_confirmed_case_list, '疫情确诊累计趋势图', city)
		draw_daily_case_figure(simplified_date_list, city_accumulated_added_cured_cases, '疫情治愈(绿)-死亡(红)累计趋势图',
							city, color='green', case_number_list2=city_accumulated_added_dead_cases, color2='red')




